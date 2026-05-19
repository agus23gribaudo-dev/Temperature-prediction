# Libreria para interpolar fillmissing
# Porque la prediccion cuando acumula muchos datos decae la temperatura o predice siempre a menor temperatura?
# Porque no re entrena

# Extracción de datos
import urllib.request # Para realizar solicitudes HTTP
import re 
import html 
# Matemáticas
import math
import numpy as np
from scipy.interpolate import interp1d 
# Tiempo
from datetime import datetime, timedelta
import time
# Memoria y archivo
import os
from collections import deque
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# Gráficos
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib
matplotlib.use("Agg")

#  PARÁMETROS  (ajustá estos valores a tu gusto)
URL       = "http://galileo4.unl.edu.ar/estacion/datos.txt"
XLSX_FILE = r"C:\Users\HP\OneDrive\Desktop\...\historial_galileo.xlsx"

POLL_INTERVAL = 60     # Intervalo de sondeo en segundos
MIN_TRAIN_PTS = 288    # Mínimo de puntos para entrenamiento
N_RETRAIN     = 2000   # Número de reentrenamientos
Iter_init     = 4000   # Iteraciones iniciales
Iter_retrain  = 2000   # Iteraciones de reentrenamiento
Lr            = 0.05   # Tasa de aprendizaje
ERR_WINDOW    = 20     # Ventana de error
RMSE_UMBRAL   = 1.0    # Umbral de RMSE
N_PASOS       = 12     # Número de pasos a predecir (5, 10, 15, 20 min, etc.)
N_ENTRADAS    = 6      # Cuántos valores pasados usa la neurona como entrada (6 × 5 min = 30 min de contexto)
HIST_PLOT     = 36     # cuántos datos reales mostrar en el gráfico (36 × 5 min = 3 horas)

PASO_MIN      = 5      # intervalo esperado entre mediciones en minutos
MAX_INTERP    = 12     # máximo de pasos faltantes a interpolar
                       # si el hueco es mayor, se deja como está (demasiado incierto)

#  LECTURA DE LA ESTACIÓN
MESES_ES = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12
}

def leer_estacion():
    try:
        req = urllib.request.Request(URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            texto = html.unescape(resp.read().decode("utf-8", errors="replace"))
    except Exception as e:
        print(f"  [WARN] Error de red: {e}")
        return None, None

    m_temp = re.search(r"T:\s*([\-\d.]+)\s*°C", texto)
    if not m_temp:
        print(f"  [WARN] Temperatura no encontrada: {texto[:100]}")
        return None, None
    temperatura = float(m_temp.group(1))

    m_hora = re.search(r"Hora:\s*(\d{1,2}:\d{2})\s+on\s+(\d{1,2})\s+(\w+)\s+(\d{4})", texto)
    if m_hora:
        hhmm, dia, mes_str, anio = m_hora.groups()
        mes = MESES_ES.get(mes_str.lower(), 1)
        dt = datetime(int(anio), mes, int(dia), int(hhmm.split(":")[0]), int(hhmm.split(":")[1]))
    else:
        dt = datetime.now().replace(second=0, microsecond=0)
    return dt, temperatura

#  INTERPOLACIÓN DE DATOS FALTANTES
def interpolar_faltantes(tiempos, temps):
    """
    Detecta huecos en la serie temporal y rellena los datos faltantes.

    - Huecos de 1 paso   → interpolación lineal
    - Huecos de 2-3 pasos → interpolación cuadrática (si hay ≥3 puntos conocidos)
    - Huecos de 4 a MAX_INTERP pasos → interpolación lineal (más conservadora)
    - Huecos > MAX_INTERP → se omiten (demasiado inciertos)

    Devuelve (tiempos_completos, temps_completos, n_interpolados)
    """
    if len(tiempos) < 2:
        return tiempos, temps, 0

    paso = timedelta(minutes=PASO_MIN)
    t_out    = [tiempos[0]]
    v_out    = [temps[0]]
    n_interp = 0

    for i in range(1, len(tiempos)):
        dt_gap = tiempos[i] - tiempos[i - 1]  # Calcular la diferencia de tiempo
        pasos_faltantes = round(dt_gap.total_seconds() / 60 / PASO_MIN) - 1

        if pasos_faltantes <= 0:
            t_out.append(tiempos[i])
            v_out.append(temps[i])
            continue

        if pasos_faltantes > MAX_INTERP:
            print(f"  [INTERP] Hueco de {pasos_faltantes} pasos entre "
                  f"{tiempos[i-1].strftime('%H:%M')} y {tiempos[i].strftime('%H:%M')} "
                  f"— demasiado grande, se omite.")
            t_out.append(tiempos[i])
            v_out.append(temps[i])
            continue

        # Elegir puntos conocidos e intentar cuadrático si hay suficientes
        # Siempre usamos al menos el punto anterior y el siguiente como anclas
        idx_izq = max(0, i - 2)
        t_conocidos = list(tiempos[idx_izq:i]) + [tiempos[i]]
        v_conocidos  = list(temps[idx_izq:i])  + [temps[i]]

        # Decidir método según puntos disponibles y tamaño del hueco
        if len(t_conocidos) >= 3 and pasos_faltantes <= 3:
            metodo = "quadratic"
        else:
            metodo = "linear"
            # Para lineal solo necesitamos los dos extremos
            t_conocidos = [tiempos[i - 1], tiempos[i]]
            v_conocidos  = [temps[i - 1],  temps[i]]

        # Convertir tiempos a segundos para interpolar numéricamente
        t0    = t_conocidos[0]
        t_seg = [(t - t0).total_seconds() for t in t_conocidos]

        try:
            interp_fn = interp1d(t_seg, v_conocidos, kind=metodo,
                                 fill_value="extrapolate")
            for k in range(1, pasos_faltantes + 1):
                t_nuevo = tiempos[i - 1] + paso * k
                v_nuevo = float(interp_fn((t_nuevo - t0).total_seconds()))
                t_out.append(t_nuevo)
                v_out.append(round(v_nuevo, 2))
                n_interp += 1
        except Exception: # Si falla cualquier método, caer a lineal simple
            t0    = tiempos[i - 1]
            t1    = tiempos[i]
            v0    = temps[i - 1]
            v1    = temps[i]
            total = (t1 - t0).total_seconds()
            for k in range(1, pasos_faltantes + 1):
                t_nuevo = t0 + paso * k
                fraccion = (t_nuevo - t0).total_seconds() / total
                v_nuevo  = round(v0 + fraccion * (v1 - v0), 2)
                t_out.append(t_nuevo)
                v_out.append(v_nuevo)
                n_interp += 1

        t_out.append(tiempos[i])
        v_out.append(temps[i])

    if n_interp > 0:
        print(f"  [INTERP] Se interpolaron {n_interp} dato(s) faltante(s).")

    return t_out, v_out, n_interp

#  EXCEL — estilos y estructura
CABECERA = ["timestamp", "temp_real_C",
            "pred_5min",  "pred_10min", "pred_15min", "pred_20min",
            "pred_25min", "pred_30min", "pred_35min", "pred_40min",
            "pred_45min", "pred_50min", "pred_55min", "pred_60min",
            "error_C", "rmse_ventana", "mae_ventana", "reentrenado"]

_thin   = Side(style="thin", color="CCCCCC")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _estilo_cabecera(cell):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", fgColor="2F4F8F")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border

def _estilo_dato(cell, col, fila_par):
    cell.font   = Font(name="Arial", size=10)
    cell.border = _border
    if col in (3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14):
        bg = "C8DEF0" if fila_par else "D6E4F0"
    elif col in (15, 16, 17, 18):
        bg = "D4EDD4" if fila_par else "EAF4EA"
    else:
        bg = "F7F9FC" if fila_par else "FFFFFF"
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left" if col == 1 else "center")

def _crear_libro():
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"
    ws.append(CABECERA)
    for col, cell in enumerate(ws[1], 1):
        _estilo_cabecera(cell)
    for col, ancho in enumerate([22,12,10,10,10,10,10,10,10,10,10,10,10,10,10,12,12,12], 1):
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    wb.save(XLSX_FILE)
    print(f"  [INFO] Nuevo archivo Excel creado: {XLSX_FILE}")

#  CARGAR HISTORIAL DESDE EXCEL
def cargar_xlsx():
    tiempos, temps = [], []
    if not os.path.exists(XLSX_FILE):
        return tiempos, temps
    wb = load_workbook(XLSX_FILE, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            ts, temp = row[0], row[1]
            if ts is None or temp is None:
                continue
            if isinstance(ts, str):
                ts = datetime.fromisoformat(ts)
            tiempos.append(ts)
            temps.append(float(temp))
        except Exception:
            pass
    wb.close()
    return tiempos, temps

#  GUARDAR FILA EN EXCEL
def guardar_fila(dt, temp_real, preds, error, rmse, mae, reentrenado):
    fila = [
        dt.strftime("%Y-%m-%d %H:%M:%S"),
        round(temp_real, 2),
        *[round(p, 2) if not math.isnan(p) else None for p in preds],
        round(error, 2) if not math.isnan(error) else None,
        round(rmse, 4),
        round(mae, 4),
        int(reentrenado),
    ]
    for intento in range(1, 6):
        try:
            if not os.path.exists(XLSX_FILE):
                _crear_libro()
            wb = load_workbook(XLSX_FILE)
            ws = wb.active
            next_row = ws.max_row + 1
            ws.append(fila)
            par = next_row % 2 == 0
            for col, cell in enumerate(ws[next_row], 1):
                _estilo_dato(cell, col, par)
            wb.save(XLSX_FILE)
            wb.close()
            return
        except PermissionError:
            print(f"  [WARN] Excel bloqueado (intento {intento}/5) — ¿está abierto? Reintentando en 5 s…")
            time.sleep(5)
    print("  [ERROR] No se pudo guardar tras 5 intentos. Cerrá el Excel y continuá.")

#  ENTRENAMIENTO
def entrenar(temps_array, Iter, w_init=None, b_init=0.0, verbose=True):
    """
    Entrena la neurona usando DIFERENCIAS de temperatura como entradas.
    Cada fila de X es una ventana [ΔT(t-N+2), ..., ΔT(t)] (N_ENTRADAS deltas).
    El target y es ΔT(t+1) = T(t+1) - T(t).
    Para predecir la temperatura real: T(t+1) = T(t) + ΔT(t+1).
    Devuelve (w, b, d_min, d_max).
    """
    temp   = np.array(temps_array, dtype=float)
    deltas = np.diff(temp)   # ΔT(i) = T(i+1) - T(i), longitud = len(temp)-1

    # Construir ventanas de N_ENTRADAS deltas consecutivos
    # X[i] = [Δ(i), Δ(i+1), ..., Δ(i+N_ENTRADAS-1)]
    # y[i] = Δ(i+N_ENTRADAS)
    n = len(deltas) - N_ENTRADAS
    X = np.array([deltas[i:i+N_ENTRADAS] for i in range(n)])
    y = deltas[N_ENTRADAS:]

    # Normalización MinMax sobre todos los deltas
    d_min, d_max = deltas.min(), deltas.max()
    rango = d_max - d_min if d_max != d_min else 1.0

    X_norm = (X - d_min) / rango
    y_norm = (y - d_min) / rango

    # Inicializar pesos
    w = w_init if w_init is not None else np.zeros(N_ENTRADAS)
    b = b_init

    if verbose:
        print(f"  Entrenando neurona: {Iter} épocas  /  lr = {Lr}  /  entradas = {N_ENTRADAS} deltas")
        print(f"  Rango deltas: [{d_min:.3f}, {d_max:.3f}] °C")
        print(f"  Pesos iniciales — w: {np.round(w, 4)}  b: {b:.6f}\n")

    for epoch in range(Iter):
        y_pred = X_norm @ w + b
        error  = y_pred - y_norm
        loss   = np.mean(error ** 2)
        dw     = (2 / n) * X_norm.T @ error
        db     = (2 / n) * np.sum(error)
        w -= Lr * dw
        b -= Lr * db

        if verbose and (epoch + 1) % (Iter // 4) == 0:
            print(f"   Época {epoch+1:>5}/{Iter}  MSE: {loss:.6f}  b: {b:.4f}")

    if verbose:
        print(f"\n  Pesos finales — w: {np.round(w, 6)}")
        print(f"  Bias  final  — b: {b:.6f}")

    return w, b, d_min, d_max

#  PREDICCIÓN
def predecir(ventana_deltas, w, b, d_min, d_max):
    """
    Dado un array de N_ENTRADAS deltas recientes, predice el próximo delta
    y lo devuelve desnormalizado en °C.
    """
    rango = d_max - d_min if d_max != d_min else 1.0
    ventana_norm = (np.array(ventana_deltas) - d_min) / rango
    delta_norm   = np.dot(ventana_norm, w) + b # multiplica cada delta pasado por su peso correspondiente y suma todo
    return float(delta_norm * rango + d_min) # desnormaliza el resultado para obtener el delta en °C

def predecir_n_pasos(temps, w, b, d_min, d_max):
    """
    Predicción iterativa N_PASOS hacia adelante usando deltas.
    - Arranca con la ventana de los últimos N_ENTRADAS deltas reales.
    - Cada paso predice ΔT, lo suma a la última temperatura conocida
      para obtener la temperatura absoluta, y actualiza la ventana.
    """
    # Ventana inicial de deltas reales
    deltas_reales = np.diff(temps[-N_ENTRADAS - 1:])  # N_ENTRADAS deltas
    ventana = deque(deltas_reales, maxlen=N_ENTRADAS) # ventana deslizante de deltas

    ultima_temp  = temps[-1]
    predicciones = []

    for _ in range(N_PASOS):
        delta_pred  = predecir(list(ventana), w, b, d_min, d_max)
        temp_pred   = ultima_temp + delta_pred # predice la temperatura absoluta
        predicciones.append(temp_pred) 
        ventana.append(delta_pred)   # el delta predicho entra a la ventana
        ultima_temp = temp_pred      # la temp predicha es la base del siguiente paso

    return predicciones

#  ESTADÍSTICAS DE ERROR  (ventana deslizante)
errores_abs = deque(maxlen=ERR_WINDOW)

def actualizar_stats(pred, real):
    errores_abs.append(abs(pred - real))

def rmse_actual():
    if not errores_abs: return 0.0
    return math.sqrt(sum(e**2 for e in errores_abs) / len(errores_abs))

def mae_actual():
    if not errores_abs: return 0.0
    return sum(errores_abs) / len(errores_abs)

#  GRÁFICO
# Inicializar figura interactiva — se crea una sola vez y se reutiliza
plt.ion()
fig, ax = plt.subplots(figsize=(13, 5))
fig.patch.set_facecolor("#0F1117")
ax.set_facecolor("#1A1D27")
fig.tight_layout(pad=2.5)

def graficar(tiempos, temps, dt_actual, preds, rmse):
    """
    Actualiza el gráfico en vivo con los datos reales y las predicciones.
    Solo muestra datos continuos recientes (sin saltos de tiempo > 15 min).
    """
    ax.cla()

    # Filtrar datos continuos: descartar si hay salto > 15 min
    MAX_GAP = timedelta(minutes=15)
    t_cont, v_cont = [], []
    for i in range(len(tiempos) - 1, -1, -1):
        if t_cont and (t_cont[0] - tiempos[i]) > MAX_GAP:
            break   # encontramos un salto, cortamos acá
        t_cont.insert(0, tiempos[i])
        v_cont.insert(0, temps[i])
        if len(t_cont) >= HIST_PLOT:
            break

    if len(t_cont) < 2:
        ax.text(0.5, 0.5, "Acumulando datos…",
                ha="center", va="center", transform=ax.transAxes,
                color="#AAAAAA", fontsize=12)
        fig.canvas.draw()
        fig.canvas.flush_events()
        plt.pause(0.01)
        return

    # Datos reales 
    ax.plot(t_cont, v_cont,
            color="#4FC3F7", linewidth=1.8, marker="o",
            markersize=3, label="Temperatura real", zorder=3)

    # Predicciones
    t_preds = [dt_actual + timedelta(minutes=5 * (i + 1)) for i in range(N_PASOS)]

    ax.plot(t_preds, preds,
            color="#FF8A65", linewidth=1.8, marker="o",
            markersize=3, linestyle="--", label="Predicción", zorder=3)

    # Banda de incertidumbre (crece con el horizonte)
    incertidumbre = [rmse * (1 + 0.15 * i) for i in range(N_PASOS)]
    v_sup = [p + u for p, u in zip(preds, incertidumbre)]
    v_inf = [p - u for p, u in zip(preds, incertidumbre)]
    ax.fill_between(t_preds, v_inf, v_sup,
                    color="#FF8A65", alpha=0.18, label="Incertidumbre (±RMSE·k)")

    # Línea vertical "Ahora"
    ax.axvline(dt_actual, color="#FFFFFF", linewidth=0.8,
               linestyle=":", alpha=0.5, label="Ahora")

    # Anotaciones
    ax.annotate(f"{v_cont[-1]:.1f}°C",
                xy=(t_cont[-1], v_cont[-1]),
                xytext=(8, 8), textcoords="offset points",
                color="#4FC3F7", fontsize=9, fontweight="bold")
    ax.annotate(f"{preds[-1]:.1f}°C  (+{N_PASOS*5}min)",
                xy=(t_preds[-1], preds[-1]),
                xytext=(8, 8), textcoords="offset points",
                color="#FF8A65", fontsize=9, fontweight="bold")

    # Formato de ejes
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    ax.xaxis.set_major_locator(mdates.MinuteLocator(byminute=range(0, 60, 15)))
    fig.autofmt_xdate(rotation=30, ha="right")

    ax.tick_params(colors="#AAAAAA", labelsize=8)
    ax.spines[["top", "right"]].set_visible(False)
    ax.spines[["left", "bottom"]].set_color("#333344")

    ax.set_ylabel("Temperatura (°C)", color="#AAAAAA", fontsize=9)
    ax.set_title(
        f"Estación Galileo — {dt_actual.strftime('%d/%m/%Y %H:%M')}  |  "
        f"RMSE={rmse:.3f}°C  |  {len(t_cont)} datos reales",
        color="#DDDDDD", fontsize=10, pad=8
    )
    ax.legend(loc="upper left", fontsize=8,
              facecolor="#1A1D27", edgecolor="#333344", labelcolor="#CCCCCC")
    ax.grid(True, color="#2A2D3A", linewidth=0.5, linestyle="--")

    fig.canvas.draw()
    fig.canvas.flush_events()
    plt.pause(0.01)
    
    nombre_archivo = "galileo_prediccion_actual.png"
    
    fig.savefig(
        nombre_archivo, 
        dpi=150, 
        bbox_inches="tight", 
        facecolor=fig.get_facecolor()
    )

    fig.canvas.draw()
    fig.canvas.flush_events()
    plt.pause(0.01)

print("=" * 62)
print("  PREDICTOR ADAPTATIVO — UNA NEURONA — Estación Galileo")
print("=" * 62)

# 1. Cargar historial 
tiempos, temps = cargar_xlsx()
print(f"\nHistorial local: {len(temps)} registros  ({XLSX_FILE})")

# Interpolar datos faltantes del historial antes de entrenar
if len(tiempos) >= 2:
    tiempos, temps, n_interp = interpolar_faltantes(tiempos, temps)
    if n_interp:
        print(f"  → Historial completado: {len(temps)} registros tras interpolar {n_interp} faltantes")

ultimo_dt = tiempos[-1] if tiempos else None

# 2. Primera lectura
print(f"Conectando a {URL} …")
dt_act, t_act = leer_estacion()

if dt_act is None:
    print("  [ERROR] No se pudo leer. Reintentando en 60 s…")
    time.sleep(60)
    dt_act, t_act = leer_estacion()
    if dt_act is None:
        raise SystemExit("  [FATAL] Sin datos de la estación.")

NANS_6 = [float("nan")] * N_PASOS

if ultimo_dt is None or dt_act > ultimo_dt:
    guardar_fila(dt_act, t_act, preds=NANS_6, error=float("nan"),
                 rmse=0.0, mae=0.0, reentrenado=False)
    temps.append(t_act)
    ultimo_dt = dt_act
    print(f"  Dato inicial almacenado: {dt_act}  T={t_act} °C")
else:
    print(f"  Último dato conocido:    {dt_act}  T={t_act} °C")

# 3. Acumulación y entrenamiento inicial
if len(temps) < MIN_TRAIN_PTS:
    print(f"\n[INFO] Sólo {len(temps)} puntos (mínimo {MIN_TRAIN_PTS}).")
    print(f"Acumulando datos… (consultando cada {POLL_INTERVAL} s)\n")
    while len(temps) < MIN_TRAIN_PTS:
        time.sleep(POLL_INTERVAL)
        dt_n, t_n = leer_estacion()
        if dt_n and dt_n > ultimo_dt:
            guardar_fila(dt_n, t_n, preds=NANS_6, error=float("nan"),
                         rmse=0.0, mae=0.0, reentrenado=False)
            temps.append(t_n)
            ultimo_dt = dt_n
            print(f"  +dato ({len(temps)}/{MIN_TRAIN_PTS})  {dt_n}  T={t_n} °C")

print(f"\nEntrenamiento inicial con {len(temps)} puntos ({Iter_init} épocas)…")
w, b, d_min, d_max = entrenar(temps[-N_RETRAIN:], Iter_init, verbose=True)

# 4. Primera predicción
preds      = predecir_n_pasos(temps, w, b, d_min, d_max)
preds_prev = preds
pred_sig   = preds[0]
print(f"\n► Primeras predicciones desde T actual = {temps[-1]:.2f} °C:")
for i, p in enumerate(preds, 1):
    print(f"     +{i*5:2d} min → {p:.2f} °C")
print(f"\nBucle activo — criterio RMSE = {RMSE_UMBRAL} °C  "
      f"(ventana {ERR_WINDOW} muestras)  /  horizonte = {N_PASOS*5} min")
print("─" * 62)

# Primer gráfico tras el entrenamiento inicial
graficar(tiempos, temps, ultimo_dt, preds, rmse_actual())

# 5. BUCLE PRINCIPAL
while True:

    while True:
        time.sleep(POLL_INTERVAL)
        dt_n, t_n = leer_estacion()
        if dt_n is None:
            continue
        if dt_n > ultimo_dt:
            break
        print(f"  … esperando nuevo dato (último: {ultimo_dt.strftime('%H:%M')})", end="\r")

    error_ciclo = t_n - pred_sig
    actualizar_stats(pred_sig, t_n)
    rmse = rmse_actual()
    mae  = mae_actual()

    reentrenado = False
    if len(errores_abs) >= ERR_WINDOW // 2 and rmse >= RMSE_UMBRAL:
        n_pts = min(len(temps), N_RETRAIN)
        print(f"\n*** RMSE={rmse:.3f} >= umbral {RMSE_UMBRAL} °C"
              f" → re-entrenando con {n_pts} puntos ({Iter_retrain} épocas) ***")
        w, b, d_min, d_max = entrenar(
            temps[-N_RETRAIN:], Iter_retrain, w_init=w, b_init=b, verbose=False)
        reentrenado = True

    guardar_fila(dt_n, t_n, preds_prev, error_ciclo, rmse, mae, reentrenado)
    temps.append(t_n)
    tiempos.append(dt_n)

    # Interpolar si hubo un hueco desde el dato anterior
    gap_pasos = round((dt_n - ultimo_dt).total_seconds() / 60 / PASO_MIN) - 1
    if 0 < gap_pasos <= MAX_INTERP:
        tiempos, temps, n_interp = interpolar_faltantes(tiempos, temps)
        if n_interp:
            print(f"  [INTERP] Hueco de {gap_pasos} paso(s) rellenado entre "
                  f"{ultimo_dt.strftime('%H:%M')} y {dt_n.strftime('%H:%M')}")

    ultimo_dt = dt_n

    preds      = predecir_n_pasos(temps, w, b, d_min, d_max)
    preds_prev = preds
    pred_sig   = preds[0]

    marca = " RE-ENTRENÓ" if reentrenado else "   continúa  "
    print(f"\n{marca}  {dt_n.strftime('%d/%m/%Y %H:%M')}")
    print(f"  Real          : {t_n:.2f} °C")
    print(f"  Pred anterior : {t_n - error_ciclo:.2f} °C   →   error = {error_ciclo:+.2f} °C")
    print(f"  RMSE ventana  : {rmse:.3f} °C   /   MAE = {mae:.3f} °C   (N={len(errores_abs)})")
    print(f"  Pesos w       : {np.round(w, 4)}  /  b = {b:.6f}")
    print(f"  Predicciones futuras ({N_PASOS*5} min):")
    for i, p in enumerate(preds, 1):
        marca_paso = " ◄ referencia" if i == 1 else ""
        print(f"     +{i*5:2d} min  [{dt_n.strftime('%H:%M')} +{i*5:2d}']  →  {p:.2f} °C{marca_paso}")
    print("─" * 62)

    # Actualizar gráfico con los nuevos datos
    graficar(tiempos, temps, dt_n, preds, rmse)